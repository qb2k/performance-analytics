"""
hmlpe.py — Гибридная многоуровневая модель оценки эффективности (HMLPE)
Формула: Ei = α·Pi + β·Ri + γ·Vi + δ·Si
"""

import pandas as pd
import numpy as np
import yaml
import re


# ─────────────────────────────────────────────
# Загрузка конфигурации
# ─────────────────────────────────────────────

def load_config(path: str = "config.yaml") -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_data(metrics_path: str, users_path: str):
    metrics = pd.read_csv(metrics_path, parse_dates=["created_at", "closed_at"])
    users = pd.read_csv(users_path)
    return metrics, users


# ─────────────────────────────────────────────
# Нормализация
# ─────────────────────────────────────────────

def normalize_minmax(series: pd.Series, minimize: bool = False) -> pd.Series:
    """Min-max нормализация. При minimize=True — обратная нормализация."""
    s = series.copy().fillna(series.median())
    mn, mx = s.min(), s.max()
    if mx == mn:
        return pd.Series(0.5, index=series.index)
    norm = (s - mn) / (mx - mn)
    return (1 - norm) if minimize else norm


# ─────────────────────────────────────────────
# P-layer: процессная эффективность
# ─────────────────────────────────────────────

def _p_score(df_u: pd.DataFrame) -> dict:
    closed_count = len(df_u[df_u["state"] == "closed"])
    avg_close = df_u["time_to_close_hours"].mean() if "time_to_close_hours" in df_u else np.nan
    avg_resp = df_u["time_to_first_response_hours"].mean() if "time_to_first_response_hours" in df_u else np.nan
    pr_count = len(df_u[df_u["type"] == "pull_request"])
    return {
        "closed_count": closed_count,
        "pr_count": pr_count,
        "avg_close_time": avg_close,
        "avg_response_time": avg_resp,
    }


# ─────────────────────────────────────────────
# R-layer: ролево-функциональная эффективность
# ─────────────────────────────────────────────

def _r_score(df_u: pd.DataFrame, role: str, config: dict) -> float:
    role_key = role.lower()
    rw = config.get("role_weights", {}).get(role_key, {"pr_merged": 0.34, "issues_closed": 0.33, "commits": 0.33})

    total = max(len(df_u), 1)
    pr_count = len(df_u[df_u["type"] == "pull_request"])
    issue_count = len(df_u[df_u["type"] == "issue"])

    score = (
        rw.get("pr_merged", 0.33) * (pr_count / total) +
        rw.get("issues_closed", 0.33) * (issue_count / total)
    )
    return score


# ─────────────────────────────────────────────
# V-layer: ценностный вклад
# ─────────────────────────────────────────────

def _v_score(df_u: pd.DataFrame, config: dict) -> float:
    value_map = {k.lower(): v for k, v in config.get("value_labels", {}).items()}
    values = []
    for _, row in df_u.iterrows():
        raw = str(row.get("labels", ""))
        labels = [l.strip().lower() for l in re.split(r"[;,]", raw) if l.strip()]
        item_val = max((value_map.get(lbl, 0.0) for lbl in labels), default=0.1)
        values.append(item_val)
    return float(np.mean(values)) if values else 0.0


# ─────────────────────────────────────────────
# S-layer: социально-командный вклад
# ─────────────────────────────────────────────

def _s_score(df_u: pd.DataFrame) -> dict:
    avg_comments = df_u["comments"].mean() if "comments" in df_u else 0.0
    avg_resp = df_u["time_to_first_response_hours"].mean() if "time_to_first_response_hours" in df_u else np.nan
    return {
        "avg_comments": avg_comments,
        "avg_response_time": avg_resp,
    }


# ─────────────────────────────────────────────
# Главная функция расчёта
# ─────────────────────────────────────────────

def calculate_scores(metrics_df: pd.DataFrame, users_df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Рассчитывает интегральные оценки эффективности всех участников.
    Возвращает DataFrame с колонками: login, role, level, team, P, R, V, S, E
    и вспомогательными метриками.
    """
    weights = config.get("weights", {"alpha": 0.30, "beta": 0.25, "gamma": 0.25, "delta": 0.20})
    alpha = weights.get("alpha", 0.30)
    beta  = weights.get("beta",  0.25)
    gamma = weights.get("gamma", 0.25)
    delta = weights.get("delta", 0.20)

    rows = []

    for _, user_row in users_df.iterrows():
        login = str(user_row["login"]).strip()
        df_u = metrics_df[metrics_df["author"].str.strip().str.lower() == login.lower()].copy()

        if len(df_u) == 0:
            continue

        p = _p_score(df_u)
        r_raw = _r_score(df_u, str(user_row.get("role", "backend")), config)
        v_raw = _v_score(df_u, config)
        s = _s_score(df_u)

        rows.append({
            "login":              login,
            "role":               user_row.get("role", ""),
            "level":              user_row.get("level", ""),
            "team":               user_row.get("team", ""),
            # сырые P-метрики
            "closed_count":       p["closed_count"],
            "pr_count":           p["pr_count"],
            "avg_close_time":     p["avg_close_time"],
            "avg_response_time":  p["avg_response_time"],
            # сырые V / R / S
            "v_raw":              v_raw,
            "r_raw":              r_raw,
            "avg_comments":       s["avg_comments"],
            "s_response":         s["avg_response_time"],
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # ── Нормализация ──────────────────────────────────────────────

    df["P"] = (
        normalize_minmax(df["closed_count"])                            * 0.50 +
        normalize_minmax(df["avg_close_time"],    minimize=True)        * 0.30 +
        normalize_minmax(df["avg_response_time"], minimize=True)        * 0.20
    )

    df["R"] = normalize_minmax(df["r_raw"])

    df["V"] = normalize_minmax(df["v_raw"])

    df["S"] = (
        normalize_minmax(df["avg_comments"])                            * 0.60 +
        normalize_minmax(df["s_response"],        minimize=True)        * 0.40
    )

    # ── Интегральная оценка ───────────────────────────────────────
    df["E"] = alpha * df["P"] + beta * df["R"] + gamma * df["V"] + delta * df["S"]

    # Перевод в 0–100
    for col in ["P", "R", "V", "S", "E"]:
        df[col] = (df[col] * 100).round(1)

    df = df.sort_values("E", ascending=False).reset_index(drop=True)
    df.index = df.index + 1  # рейтинг с 1

    return df


# ─────────────────────────────────────────────
# Экспорт в Excel
# ─────────────────────────────────────────────

def export_excel(df: pd.DataFrame, weights: dict) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    wb = Workbook()

    # ── Лист 1: Сводная таблица ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Рейтинг"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = ["login", "role", "level", "team", "P", "R", "V", "S", "E",
            "closed_count", "pr_count", "avg_close_time", "avg_response_time", "avg_comments"]
    col_names = {
        "login": "Участник", "role": "Роль", "level": "Уровень", "team": "Команда",
        "P": "P-layer", "R": "R-layer", "V": "V-layer", "S": "S-layer", "E": "Оценка E",
        "closed_count": "Закрыто задач", "pr_count": "PR",
        "avg_close_time": "Ср. время закрытия (ч)",
        "avg_response_time": "Ср. время ответа (ч)",
        "avg_comments": "Ср. комментариев",
    }

    ws1.append(["#"] + [col_names.get(c, c) for c in cols])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    score_fills = {
        "high":   PatternFill("solid", fgColor="C6EFCE"),
        "medium": PatternFill("solid", fgColor="FFEB9C"),
        "low":    PatternFill("solid", fgColor="FFC7CE"),
    }

    for i, (_, row) in enumerate(df.reset_index().iterrows(), start=2):
        ws1.append([row.get("index", i)] + [row.get(c, "") for c in cols])
        e_val = row.get("E", 50)
        fill = score_fills["high"] if e_val >= 65 else score_fills["medium"] if e_val >= 45 else score_fills["low"]
        for cell in ws1[i]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        ws1.cell(row=i, column=cols.index("E") + 2).fill = fill

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

    # ── Лист 2: Слои по участникам ───────────────────────────────
    ws2 = wb.create_sheet("Слои HMLPE")
    ws2.append(["Участник", "P-layer", "R-layer", "V-layer", "S-layer", "Оценка E"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.reset_index().iterrows():
        ws2.append([row["login"], row["P"], row["R"], row["V"], row["S"], row["E"]])

    chart = BarChart()
    chart.type = "col"
    chart.title = "Слои HMLPE по участникам"
    chart.y_axis.title = "Оценка (0–100)"
    chart.x_axis.title = "Участник"
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=len(df) + 1)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "H2")

    # ── Лист 3: Веса конфигурации ─────────────────────────────────
    ws3 = wb.create_sheet("Конфигурация")
    ws3.append(["Параметр", "Значение"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws3.append(["α (P-layer)", weights.get("alpha")])
    ws3.append(["β (R-layer)", weights.get("beta")])
    ws3.append(["γ (V-layer)", weights.get("gamma")])
    ws3.append(["δ (S-layer)", weights.get("delta")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
